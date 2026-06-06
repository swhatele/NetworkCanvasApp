"""
TNN Network Analyzer — Connecting for Change
v5 — cleaner UI, projectable insights export
"""

import io, zipfile, xml.etree.ElementTree as ET, math, collections
import pandas as pd
import networkx as nx
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="TNN Network Analyzer", page_icon="🌐",
                   layout="wide", initial_sidebar_state="expanded")

# ── Palette ───────────────────────────────────────────────────────────────────
INDIGO="#2825BE"; INDIGO_LT="#4a47d6"; AMBER="#EB9001"; TEAL="#0C7A7A"
TEAL_LT="#0fa8a8"; TERRA="#CF4C38"; GREEN="#2E9E5B"; INK="#080818"
INK2="#0f0e2e"; INK3="#13123a"; SURFACE="#ffffff"; SURFACE2="#F5F7F6"
BORDER="#e2e5e3"; TEXT="#111827"; TEXT2="#4b5563"; TEXT3="#9ca3af"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@600;700;800;900&family=IBM+Plex+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&family=IBM+Plex+Mono:wght@400;500&display=swap');
html,body,[class*="css"]{font-family:'IBM Plex Sans',sans-serif;color:#111827;}
h1,h2,h3{font-family:'Barlow Condensed',sans-serif;letter-spacing:-0.01em;color:#080818;}
.eyebrow{font-size:10.5px;font-weight:700;letter-spacing:0.16em;text-transform:uppercase;color:#EB9001;margin-bottom:4px;font-family:'IBM Plex Sans',sans-serif;}
.metric-card{background:#F5F7F6;border-radius:8px;padding:20px 24px;border-left:3px solid #2825BE;}
.metric-card .value{font-family:'Barlow Condensed',sans-serif;font-size:44px;font-weight:800;color:#2825BE;line-height:1;}
.metric-card .label{font-family:'IBM Plex Mono',monospace;font-size:11px;color:#4b5563;margin-top:4px;}
.scorecard-row{display:flex;gap:16px;margin:16px 0;}
.scorecard{flex:1;background:#F5F7F6;border-radius:8px;padding:18px 20px;border-top:3px solid #2825BE;}
.scorecard .sc-val{font-family:'Barlow Condensed',sans-serif;font-size:36px;font-weight:800;line-height:1;margin-bottom:4px;}
.scorecard .sc-label{font-family:'IBM Plex Mono',monospace;font-size:10px;color:#4b5563;text-transform:uppercase;letter-spacing:0.08em;}
        sc_html+=f'''<div class="scorecard" style="border-top-color:{c};">
<div class="sc-val" style="color:{c};">{s}%</div>
<div class="sc-label">{attr}</div>
{"<div class="sc-sub">" + sub + "</div>" if sub else ""}
</div>'''
.finding{border-bottom:1px solid #e2e5e3;padding:20px 0;}
.finding:last-child{border-bottom:none;}
.finding .f-label{font-family:'IBM Plex Mono',monospace;font-size:10px;color:#EB9001;text-transform:uppercase;letter-spacing:0.12em;margin-bottom:6px;}
.finding .f-head{font-family:'Barlow Condensed',sans-serif;font-size:22px;font-weight:800;color:#080818;margin-bottom:6px;line-height:1.2;}
.finding .f-body{font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:#4b5563;line-height:1.7;max-width:680px;}
.finding .f-stat{font-family:'IBM Plex Mono',monospace;font-size:28px;font-weight:500;color:#2825BE;margin-bottom:4px;}
.skill-match{background:#F5F7F6;border-radius:6px;padding:12px 16px;margin-bottom:8px;border-left:2px solid #0C7A7A;}
.skill-match .sm-skill{font-family:'IBM Plex Mono',monospace;font-size:11px;color:#0C7A7A;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.06em;}
.skill-match .sm-body{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:#111827;}
.stTabs [data-baseweb="tab"]{font-family:'IBM Plex Sans',sans-serif;font-size:13px;font-weight:500;color:#4b5563;}
.stTabs [aria-selected="true"]{color:#2825BE !important;}
</style>""", unsafe_allow_html=True)

# ── Lookup tables ─────────────────────────────────────────────────────────────
FREQUENCY_LABELS={1:"Never or rarely",2:"Once or twice a month",3:"Weekly",4:"Multiple times a week",5:"Daily or near-daily"}
DEPTH_LABELS={1:"Awareness",2:"Connection",3:"Cooperation",4:"Collaboration"}
KNOWLEDGE_LABELS={5:"Not at all",6:"Somewhat",7:"Quite a bit",8:"A great deal"}
SCALE4_LABELS={1:"Not at all",2:"Somewhat",3:"Quite a bit",4:"A great deal"}
TYPE_LABELS={"Type_1":"Peer Learning / Knowledge Exchange","Type_2":"Joint Programming or Co-delivery","Type_3":"Referrals","Type_4":"Funding","Type_5":"Informal Support / Thought Partnership","Type_6":"Governance or Board-level Connection"}
RECIPROCITY_LABELS={1:"Mostly flows from them to me",2:"Roughly balanced",3:"Mostly flows from me to them",4:"I'm not sure"}
SECTOR_LABELS={1:"Education",2:"Health & Human Services",3:"Community Development & Housing",4:"Government & Public Sector",5:"Faith Communities",6:"Business & Commerce",7:"Arts, Culture & Humanities",8:"Civic & Advocacy",9:"Media & Communications",10:"Environment & Conservation",11:"Philanthropy & Funding",12:"Other"}
SHARING_LABELS={1:"Place & Built Environment",2:"Recognition",3:"Rhythm & Recurrence",4:"Mutual Obligation",5:"Common Stakes",6:"Institutional Anchors",7:"Story, Memory & Identity"}
CONFIDENTIALITY_LABELS={1:"Yes – share full responses",2:"Yes – aggregate/anonymized only",3:"No – keep private"}
GML_NS="http://graphml.graphdrawing.org/xmlns"; MAX_BARE=3

def _gml(t): return f"{{{GML_NS}}}{t}"
def lbl(val,lookup):
    if pd.isna(val): return ""
    try: return lookup.get(int(val),val)
    except: return val
def bool_to_types(row,type_cols):
    s=[TYPE_LABELS.get(c,c) for c in type_cols if str(row.get(c,False)).strip().lower() in ("true","1","yes")]
    return "; ".join(s) if s else ""
def has_attrs(row):
    for col in ["Depth","Frequency","Reciprocity","Trust","Energy","Support","Creativity","Knowledge"]:
        v=row.get(col)
        if v is not None:
            try:
                if not pd.isna(v): return True
            except: return True
    return False
def assign_question(df):
    qs,bare=[],0
    for _,r in df.iterrows():
        if has_attrs(r): qs.append("Current Connections")
        else:
            bare+=1
            qs.append("Aspired Connections" if bare<=MAX_BARE else ("Informal Influence" if bare<=MAX_BARE*2 else "Unknown"))
    df=df.copy(); df["network_question"]=qs; return df

# ── Parsers ───────────────────────────────────────────────────────────────────
def parse_ego(b):
    df=pd.read_csv(io.BytesIO(b),encoding="utf-8-sig",encoding_errors="replace")
    if df.empty: return {}
    r=df.iloc[0]
    return {"ego_uuid":r.get("networkCanvasEgoUUID",""),"case_id":r.get("networkCanvasCaseID",""),
            "name":r.get("Name",""),"geography":r.get("Geography",""),
            "sector":lbl(r.get("Sector"),SECTOR_LABELS),"sharing":lbl(r.get("Sharing"),SHARING_LABELS),
            "confidentiality":lbl(r.get("Confidentiality"),CONFIDENTIALITY_LABELS),
            "consent":str(r.get("Consent","")),"expansion":r.get("Expansion",""),
            "open":r.get("Open",""),"tenure":r.get("Tenure",""),
            "session_start":r.get("sessionStart",""),"session_finish":r.get("sessionFinish","")}

def parse_attributes(b,ego_uuid,ego_name):
    df=pd.read_csv(io.BytesIO(b),encoding="utf-8-sig",encoding_errors="replace"); df=assign_question(df)
    type_cols=[c for c in df.columns if c.startswith("Type_")]
    nodes=[]
    for _,r in df.iterrows():
        nodes.append({"ego_uuid":ego_uuid,"ego_name":ego_name,"node_uuid":r.get("networkCanvasUUID",""),
            "name":r.get("Name",""),"organization":r.get("Organization",""),
            "network_question":r.get("network_question",""),
            "frequency":lbl(r.get("Frequency"),FREQUENCY_LABELS),"depth":lbl(r.get("Depth"),DEPTH_LABELS),
            "knowledge":lbl(r.get("Knowledge"),KNOWLEDGE_LABELS),"energy":lbl(r.get("Energy"),SCALE4_LABELS),
            "support":lbl(r.get("Support"),SCALE4_LABELS),"creativity":lbl(r.get("Creativity"),SCALE4_LABELS),
            "trust":lbl(r.get("Trust"),SCALE4_LABELS),"relationship_type":bool_to_types(r,type_cols),
            "reciprocity":lbl(r.get("Reciprocity"),RECIPROCITY_LABELS)})
    return nodes

def parse_graphml(b,ego_uuid,ego_name,uuid_to_name):
    tree=ET.parse(io.BytesIO(b)); root=tree.getroot()
    key_map={k.attrib.get("id",""):k.attrib.get("attr.name","") for k in root.findall(_gml("key"))}
    graph=root.find(_gml("graph"))
    if graph is None: return []
    node_data={}
    for node in graph.findall(_gml("node")):
        nid=node.attrib.get("id","")
        props={key_map.get(d.attrib.get("key",""),d.attrib.get("key","")):d.text or "" for d in node.findall(_gml("data"))}
        node_data[nid]={"uuid":props.get("networkCanvasUUID",""),"name":props.get("Name","")}
        if props.get("networkCanvasUUID"): uuid_to_name[props["networkCanvasUUID"]]=props.get("Name","")
    edges=[]
    for edge in graph.findall(_gml("edge")):
        props={key_map.get(d.attrib.get("key",""),d.attrib.get("key","")):d.text or "" for d in edge.findall(_gml("data"))}
        sn=node_data.get(edge.attrib.get("source",""),{}); tn=node_data.get(edge.attrib.get("target",""),{})
        edges.append({"From":sn.get("name",""),"To":tn.get("name","")})
    return edges

def extract_zips(uploaded_zips):
    buckets={}
    for f in uploaded_zips:
        raw=f.read()
        if not zipfile.is_zipfile(io.BytesIO(raw)): st.warning(f"⚠️ `{f.name}` is not a valid zip — skipped."); continue
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for zname in zf.namelist():
                if zname.endswith("/") or "__MACOSX" in zname.split("/"): continue
                basename=zname.split("/")[-1]
                if not basename or basename.startswith("._"): continue
                fb=zf.read(zname)
                if basename.endswith("_ego.csv"): buckets.setdefault(basename.replace("_ego.csv",""),{})["ego"]=fb
                elif basename.endswith("_attributeList_Person.csv"): buckets.setdefault(basename.replace("_attributeList_Person.csv",""),{})["attrs"]=fb
                elif basename.endswith("_edgeList.csv"): buckets.setdefault(basename.replace("_edgeList.csv",""),{})["edges_csv"]=fb
                elif basename.endswith(".graphml"): buckets.setdefault(basename.replace(".graphml",""),{})["graphml"]=fb
    return buckets

def process_buckets(buckets):
    all_egos,all_edges,warnings=[],[],[]
    uuid_to_name={}
    for sid,files in buckets.items():
        if "ego" not in files: warnings.append(f"No ego file for `{sid}` — skipped."); continue
        ego=parse_ego(files["ego"])
        ego_uuid=ego.get("ego_uuid",sid); ego_name=ego.get("name") or ego.get("case_id") or sid
        uuid_to_name[ego_uuid]=ego_name; all_egos.append(ego)
        nodes=[]
        if "attrs" in files: nodes=parse_attributes(files["attrs"],ego_uuid,ego_name)
        for n in nodes:
            if n["network_question"]!="Current Connections": continue
            all_edges.append({"From":ego_name,"To":n["name"],"To Organization":n["organization"],
                "Frequency of Interaction":n["frequency"],"Depth of Connection":n["depth"],
                "Knowledge of Their Work":n["knowledge"],"Energy":n["energy"],"Support":n["support"],
                "Creativity":n["creativity"],"Trust":n["trust"],"Relationship Type(s)":n["relationship_type"],
                "Reciprocity":n["reciprocity"]})
        if "graphml" in files: parse_graphml(files["graphml"],ego_uuid,ego_name,uuid_to_name)
    ego_df=pd.DataFrame(all_egos).drop(columns=["ego_uuid","case_id"],errors="ignore").rename(columns={
        "name":"Name","geography":"Geography","sector":"Sector",
        "sharing":"Dimension Most Excited to Share","confidentiality":"Confidentiality Preference",
        "consent":"Consent","expansion":"Expansion / Other Members","open":"Open to New Members?",
        "tenure":"Tenure","session_start":"Session Start","session_finish":"Session Finish"})
    edge_df=pd.DataFrame(all_edges) if all_edges else pd.DataFrame()
    return ego_df,edge_df,warnings

def parse_presurvey(uploaded_file):
    df=pd.read_excel(uploaded_file,sheet_name="Form Responses 1")
    df=df.rename(columns={
        "Name":"name","Organization Name":"organization","Sector":"sector_raw",
        "Skills you personally can offer (to help others in this network)":"skills_offered",
        "Skills you (or your org) need help with (to accomplish your mission)":"skills_needed",
        "Who is your org's primary target audience? (i.e., who are you serving? The more specific, the better)":"target_audience",
        "How would you describe the most challenging issue you (or your org) face? ":"challenge",
        "How would you describe your context? [Historic / Stagnant]":"ctx_historic",
        "How would you describe your context? [Transitional]":"ctx_transitional",
        "How would you describe your context? [Affluent ]":"ctx_affluent",
        "How would you describe your context? [University / Transient]":"ctx_university",
    })
    def parse_contexts(row):
        found=[]
        for col,lbl_ in [("ctx_historic","Historic/Stagnant"),("ctx_transitional","Transitional"),
                          ("ctx_affluent","Affluent"),("ctx_university","University/Transient")]:
            if pd.notna(row.get(col,"")): found.append(lbl_)
        return ", ".join(found) if found else ""
    df["neighborhood_contexts"]=df.apply(parse_contexts,axis=1)
    df["name_clean"]=df["name"].str.strip().str.lower()
    return df[["name","name_clean","organization","sector_raw","skills_offered","skills_needed","challenge","neighborhood_contexts","target_audience"]]

def merge_presurvey(ego_df,pre_df):
    if ego_df.empty or pre_df.empty: return ego_df
    ego_df=ego_df.copy()
    ego_df["name_clean"]=ego_df["Name"].str.strip().str.lower()
    merged=ego_df.merge(pre_df[["name_clean","skills_offered","skills_needed","challenge","neighborhood_contexts","sector_raw"]],
                        on="name_clean",how="left").drop(columns=["name_clean"])
    return merged

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(ego_df,edge_df):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        ego_df.to_excel(w,sheet_name="Nodes",index=False)
        if not edge_df.empty: edge_df.to_excel(w,sheet_name="Edges",index=False)
    buf.seek(0); wb=load_workbook(buf)
    hf=PatternFill("solid",start_color="2825BE"); hfont=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    bf=Font(name="Arial",size=10); alt=PatternFill("solid",start_color="F5F7F6")
    ca=Alignment(horizontal="center",vertical="top",wrap_text=True)
    la=Alignment(horizontal="left",vertical="top",wrap_text=True)
    tb=Border(bottom=Side(style="thin",color="e2e5e3"),right=Side(style="thin",color="e2e5e3"))
    for ws in wb.worksheets:
        for cell in ws[1]: cell.font=hfont; cell.fill=hf; cell.alignment=ca; cell.border=tb
        ws.freeze_panes="A2"; ws.row_dimensions[1].height=30
        for ri,row in enumerate(ws.iter_rows(min_row=2),start=2):
            fill=alt if ri%2==0 else None
            for cell in row:
                cell.font=bf; cell.alignment=la; cell.border=tb
                if fill: cell.fill=fill
        for cc in ws.columns:
            ml=max((len(str(c.value)) if c.value else 0) for c in cc)
            ws.column_dimensions[get_column_letter(cc[0].column)].width=min(ml+4,50)
    out=io.BytesIO(); wb.save(out); out.seek(0); return out

# ── Graph ─────────────────────────────────────────────────────────────────────
def build_graph(edge_df):
    G=nx.DiGraph()
    if edge_df.empty: return G
    for _,r in edge_df.iterrows():
        frm=str(r["From"]).strip(); to=str(r["To"]).strip()
        if not frm or not to or frm in ("nan","None") or to in ("nan","None"): continue
        if frm==to: continue  # skip self-loops
        if G.has_edge(frm,to): G[frm][to]["weight"]=G[frm][to].get("weight",1)+1
        else: G.add_edge(frm,to,weight=1)
    return G

# ── Network figure (light bg, visible edges) ──────────────────────────────────
def make_network_fig(G,color_by="degree",ego_names=None,node_meta=None,show_labels=True):
    if len(G.nodes)==0: return go.Figure()
    UG=G.to_undirected()
    pos=nx.spring_layout(UG,seed=42,k=2.8/math.sqrt(max(len(G.nodes),1)))
    degree=dict(G.degree()); bw=nx.betweenness_centrality(G); indeg=dict(G.in_degree())
    ego_set=set(ego_names or []); max_deg=max(degree.values()) if degree else 1

    def get_color(n):
        if color_by=="degree":
            v=degree.get(n,0); t=0.3+0.7*(v/max(max_deg,1))
            return f"rgba(40,37,190,{t:.2f})"
        elif color_by=="betweenness":
            v=bw.get(n,0); mv=max(bw.values()) if bw else 1; t=0.3+0.7*(v/max(mv,0.001))
            return f"rgba(12,122,122,{t:.2f})"
        elif color_by=="in_degree":
            v=indeg.get(n,0); mi=max(indeg.values()) if indeg else 1; t=0.3+0.7*(v/max(mi,1))
            return f"rgba(235,144,1,{t:.2f})"
        elif color_by=="survey_taker": return AMBER if n in ego_set else INDIGO
        elif color_by=="isolated": return TERRA if indeg.get(n,0)==0 and n not in ego_set else INDIGO
        return INDIGO

    node_colors=[get_color(n) for n in G.nodes]
    node_sizes=[12+30*(degree.get(n,0)/max(max_deg,1)) for n in G.nodes]

    def hover(n):
        lines=[f"<b>{n}</b>","",f"Connections: {degree.get(n,0)}",f"Times named: {indeg.get(n,0)}",f"Betweenness: {bw.get(n,0):.3f}"]
        if node_meta and n in node_meta:
            m=node_meta[n]
            if m.get("challenge"): lines+=["",f"<i>Challenge:</i> {str(m['challenge'])[:120]}"]
            if m.get("skills_offered"): lines+=[f"<i>Offers:</i> {str(m['skills_offered'])[:100]}"]
        return "<br>".join(lines)

    ex,ey=[],[]
    for u,v in G.edges():
        x0,y0=pos[u]; x1,y1=pos[v]
        ex+=[x0,x1,None]; ey+=[y0,y1,None]

    edge_trace=go.Scatter(x=ex,y=ey,mode="lines",
        line=dict(width=1.2,color="rgba(40,37,190,0.25)"),hoverinfo="none",showlegend=False)
    mode="markers+text" if show_labels else "markers"
    node_trace=go.Scatter(
        x=[pos[n][0] for n in G.nodes],y=[pos[n][1] for n in G.nodes],
        mode=mode,text=[n for n in G.nodes],textposition="top center",
        textfont=dict(family="IBM Plex Mono",size=9,color=TEXT2),
        hovertext=[hover(n) for n in G.nodes],hoverinfo="text",
        marker=dict(size=node_sizes,color=node_colors,line=dict(width=2,color="white")),
        showlegend=False)

    fig=go.Figure(data=[edge_trace,node_trace])
    fig.update_layout(
        paper_bgcolor=SURFACE2,plot_bgcolor=SURFACE2,
        margin=dict(l=16,r=16,t=16,b=16),
        xaxis=dict(showgrid=False,zeroline=False,showticklabels=False,showline=False),
        yaxis=dict(showgrid=False,zeroline=False,showticklabels=False,showline=False),
        height=560,font=dict(family="IBM Plex Sans"),
        hoverlabel=dict(bgcolor=INK2,font_color="white",font_family="IBM Plex Sans",align="left"))
    return fig

# ── Analytics helpers ─────────────────────────────────────────────────────────
def pct_high(series):
    counts=series.value_counts()
    high=counts.get("A great deal",0)+counts.get("Quite a bit",0)
    return round(100*high/max(counts.sum(),1))

def pct_label(val,series,order):
    counts=series.value_counts()
    total=counts.sum()
    return round(100*counts.get(val,0)/max(total,1))

def parse_skills(s):
    if pd.isna(s) or not str(s).strip(): return []
    return [x.strip() for x in str(s).split(",") if x.strip()]

def build_skill_svg(show_m):
    """Build an SVG bipartite matchmaking diagram without nested f-strings."""
    skill_colors = [INDIGO, TEAL, AMBER, GREEN, TERRA, INDIGO_LT]
    left_names = sorted(set(n for m in show_m for n in m["offerers"]))[:12]
    right_names = sorted(set(n for m in show_m for n in m["needers"]))[:12]
    svg_h = max(400, max(len(left_names), len(right_names)) * 36 + 80)
    svg_w = 700
    row_h = (svg_h - 80) / max(max(len(left_names),1), max(len(right_names),1))
    left_y  = {n: 40 + i*row_h + row_h/2 for i,n in enumerate(left_names)}
    right_y = {n: 40 + i*row_h + row_h/2 for i,n in enumerate(right_names)}
    parts = []
    parts.append('<svg xmlns="http://www.w3.org/2000/svg" width="100%" viewBox="0 0 ' + str(svg_w) + ' ' + str(svg_h) + '" style="background:' + SURFACE2 + ';border-radius:6px;">')
    parts.append('<text x="180" y="20" text-anchor="middle" font-family="IBM Plex Mono" font-size="9" fill="' + TEAL + '" letter-spacing="2">OFFERS</text>')
    parts.append('<text x="' + str(svg_w-180) + '" y="20" text-anchor="middle" font-family="IBM Plex Mono" font-size="9" fill="' + AMBER + '" letter-spacing="2">NEEDS</text>')
    # Lines
    for si, m in enumerate(show_m):
        color = skill_colors[si % len(skill_colors)]
        for o in m["offerers"]:
            if o in left_y:
                for n in m["needers"]:
                    if n in right_y and o != n:
                        lx=180; ly=left_y[o]; rx=svg_w-180; ry=right_y[n]; mx=svg_w/2
                        parts.append('<path d="M' + str(lx) + ',' + str(ly) + ' C' + str(mx) + ',' + str(ly) + ' ' + str(mx) + ',' + str(ry) + ' ' + str(rx) + ',' + str(ry) + '" stroke="' + color + '" stroke-width="1.5" fill="none" opacity="0.4"/>')
    # Left nodes
    for n in left_names:
        cy = str(left_y[n]); cy4 = str(left_y[n]+4)
        parts.append('<circle cx="180" cy="' + cy + '" r="5" fill="' + INDIGO + '"/>')
        parts.append('<text x="170" y="' + cy4 + '" text-anchor="end" font-family="IBM Plex Mono" font-size="10" fill="' + TEXT2 + '">' + n[:22] + '</text>')
    # Right nodes
    for n in right_names:
        cy = str(right_y[n]); cy4 = str(right_y[n]+4); cx = str(svg_w-180); cx2 = str(svg_w-170)
        parts.append('<circle cx="' + cx + '" cy="' + cy + '" r="5" fill="' + AMBER + '"/>')
        parts.append('<text x="' + cx2 + '" y="' + cy4 + '" text-anchor="start" font-family="IBM Plex Mono" font-size="10" fill="' + TEXT2 + '">' + n[:22] + '</text>')
    # Legend
    for si, m in enumerate(show_m):
        x1=str(20+si*110); y1=str(svg_h-28); x2=str(32+si*110); y2=str(svg_h-21)
        parts.append('<rect x="' + x1 + '" y="' + y1 + '" width="8" height="8" fill="' + skill_colors[si%len(skill_colors)] + '"/>')
        parts.append('<text x="' + x2 + '" y="' + y2 + '" font-family="IBM Plex Sans" font-size="9" fill="' + TEXT3 + '">' + m["skill"][:14] + '</text>')
    parts.append('</svg>')
    return chr(10).join(parts)



def build_skill_match(pre_df):
    off=collections.defaultdict(list); need=collections.defaultdict(list)
    for _,r in pre_df.iterrows():
        n=r.get("name","")
        for sk in parse_skills(r.get("skills_offered","")): off[sk].append(n)
        for sk in parse_skills(r.get("skills_needed","")): need[sk].append(n)
    all_sk=set(off)|set(need)
    return sorted([{"skill":sk,"offered":len(off.get(sk,[])),"needed":len(need.get(sk,[])),
                    "gap":len(need.get(sk,[]))-len(off.get(sk,[])),
                    "offerers":off.get(sk,[]),"needers":need.get(sk,[])} for sk in all_sk],
                  key=lambda x:-x["needed"])

# ── Network analytics summary (for both app + HTML export) ────────────────────
def compute_analytics(ego_df,edge_df,G,pre_df=None):
    """Returns a structured dict of key findings."""
    a={}
    if edge_df.empty: return a
    ego_set=set(ego_df["Name"].dropna().str.strip()) if "Name" in ego_df.columns else set()
    indeg=dict(G.in_degree()); bw=nx.betweenness_centrality(G)

    a["n_respondents"]=len(ego_df)
    a["n_nodes"]=len(G.nodes)
    a["n_edges"]=len(G.edges)
    a["density"]=round(nx.density(G),3)

    # Depth
    if "Depth of Connection" in edge_df.columns:
        dc=edge_df["Depth of Connection"].value_counts(normalize=True)*100
        a["pct_deep"]=round(dc.get("Collaboration",0)+dc.get("Cooperation",0))
        a["pct_shallow"]=round(dc.get("Awareness",0)+dc.get("Connection",0))
        depth_counts=edge_df["Depth of Connection"].value_counts()
        a["depth_counts"]={k:int(v) for k,v in depth_counts.items()}

    # Relational quality
    for attr in ["Energy","Support","Creativity","Trust","Knowledge"]:
        if attr in edge_df.columns:
            a[f"pct_{attr.lower()}"]=pct_high(edge_df[attr])

    # Low-trust edges
    if "Trust" in edge_df.columns:
        low_trust=(edge_df["Trust"].isin(["Not at all","Somewhat"])).sum()
        a["n_low_trust"]=int(low_trust)
        a["pct_low_trust"]=round(100*low_trust/max(len(edge_df),1))

    # Low-energy edges
    if "Energy" in edge_df.columns:
        low_e=(edge_df["Energy"].isin(["Not at all","Somewhat"])).sum()
        a["n_low_energy"]=int(low_e)
        a["pct_low_energy"]=round(100*low_e/max(len(edge_df),1))

    # Reciprocity
    if "Reciprocity" in edge_df.columns:
        a["pct_balanced"]=round(100*(edge_df["Reciprocity"]=="Roughly balanced").sum()/max(len(edge_df),1))
        a["pct_one_way"]=round(100*(edge_df["Reciprocity"].isin(["Mostly flows from them to me","Mostly flows from me to them"])).sum()/max(len(edge_df),1))

    # Frequency
    if "Frequency of Interaction" in edge_df.columns:
        fc=edge_df["Frequency of Interaction"].value_counts(normalize=True)*100
        a["pct_frequent"]=round(fc.get("Weekly",0)+fc.get("Multiple times a week",0)+fc.get("Daily or near-daily",0))
        a["pct_infrequent"]=round(fc.get("Never or rarely",0)+fc.get("Once or twice a month",0))

    # Relationship types
    if "Relationship Type(s)" in edge_df.columns:
        all_t=[]
        for v in edge_df["Relationship Type(s)"].dropna():
            all_t.extend([t.strip() for t in str(v).split(";") if t.strip()])
        if all_t:
            tc=collections.Counter(all_t)
            a["top_rel_type"]=tc.most_common(1)[0][0]
            a["rel_type_counts"]=dict(tc.most_common(6))

    # Isolation
    survey_lonely=[n for n in ego_set if indeg.get(n,0)==0]
    periphery=[n for n in G.nodes if indeg.get(n,0)==0 and n not in ego_set]
    a["n_survey_isolated"]=len(survey_lonely)
    a["n_periphery"]=len(periphery)

    # Bridges
    top_bw=sorted(bw.items(),key=lambda x:-x[1])
    if top_bw: a["top_bridge"]=top_bw[0][0]; a["top_bridge_score"]=round(top_bw[0][1],3)

    # Most named
    if "To" in edge_df.columns:
        top=edge_df["To"].value_counts().head(5)
        a["most_named"]=[(n,int(c)) for n,c in top.items()]

    # Geographic spread
    if "Geography" in ego_df.columns:
        geos=ego_df["Geography"].dropna()
        a["n_geographies"]=int(geos.nunique())
        a["geo_counts"]=dict(geos.value_counts())

    # Communities
    try:
        UG_clean=nx.Graph()
        for u,v in G.to_undirected().edges():
            if isinstance(u,str) and isinstance(v,str) and u!=v:
                UG_clean.add_edge(u,v)
        for n in G.nodes():
            if isinstance(n,str): UG_clean.add_node(n)
        comms=nx.community.louvain_communities(UG_clean,seed=42)
        a["n_communities"]=len(comms)
        a["communities"]=[sorted(list(c)) for c in sorted(comms,key=len,reverse=True)]
    except: a["n_communities"]=1; a["communities"]=[]

    # Skills
    if pre_df is not None and not pre_df.empty:
        matches=build_skill_match(pre_df)
        a["skill_matches"]=matches
        gaps=[m for m in matches if m["gap"]>0]
        a["top_skill_gaps"]=gaps[:3]
        surplus=[m for m in matches if m["gap"]<0]
        a["top_skill_surplus"]=surplus[:2]

    return a

# ── HTML insights generator ───────────────────────────────────────────────────
def build_insights_html(a, ego_df, edge_df):
    """Build a projectable, fully anonymized HTML document for room display."""

    # ── Pre-compute all values ─────────────────────────────────────────────
    n_resp         = a.get("n_respondents", 0)
    n_nodes        = a.get("n_nodes", 0)
    n_edges        = a.get("n_edges", 0)
    density        = a.get("density", 0)
    pct_deep       = a.get("pct_deep", 0)
    pct_trust      = a.get("pct_trust", 0)
    pct_energy     = a.get("pct_energy", 0)
    pct_creativity = a.get("pct_creativity", 0)
    pct_support    = a.get("pct_support", 0)
    pct_low_trust  = a.get("pct_low_trust", 0)
    pct_low_energy = a.get("pct_low_energy", 0)
    pct_balanced   = a.get("pct_balanced", 0)
    pct_frequent   = a.get("pct_frequent", 0)
    pct_infrequent = a.get("pct_infrequent", 0)
    n_communities  = a.get("n_communities", 0)
    n_geographies  = a.get("n_geographies", 0)
    n_isolated     = a.get("n_survey_isolated", 0)
    n_periphery    = a.get("n_periphery", 0)
    top_rel        = a.get("top_rel_type", "Peer Learning / Knowledge Exchange")
    communities    = a.get("communities", [])

    # ── Depth bar segments ─────────────────────────────────────────────────
    dc = a.get("depth_counts", {})
    def _seg(label, bg):
        v = dc.get(label, 0)
        if not v: return ""
        return '<div class="depth-seg" style="flex:' + str(v) + ';background:' + bg + ';">' + str(v) + '</div>'
    dc_awareness_html     = _seg("Awareness",    "#13123a")
    dc_connection_html    = _seg("Connection",   "#1e1c4a")
    dc_cooperation_html   = _seg("Cooperation",  "#2825BE")
    dc_collaboration_html = _seg("Collaboration","#4a47d6")

    # ── Sectors ────────────────────────────────────────────────────────────
    sector_html = ""
    if "Sector" in ego_df.columns:
        for s, c in ego_df["Sector"].value_counts().items():
            sector_html += '<div class="chip">' + str(s) + ' <span class="chip-n">' + str(c) + '</span></div>'

    # ── SVG US Map ─────────────────────────────────────────────────────────
    STATE_MAP = {
        "atlanta":"GA","georgia":"GA","ga":"GA","south atlanta":"GA",
        "chicago":"IL","illinois":"IL","il":"IL",
        "minneapolis":"MN","minnesota":"MN","twin cities":"MN","mn":"MN",
        "kansas":"KS","kansas city":"KS","ks":"KS",
        "oregon":"OR","portland":"OR","or":"OR",
        "alabama":"AL","birmingham":"AL","al":"AL",
        "new jersey":"NJ","nj":"NJ","eastern pa":"PA","pennsylvania":"PA","pa":"PA",
        "north carolina":"NC","nc":"NC","asheville":"NC","charlotte":"NC","black mountain":"NC",
        "new york":"NY","ny":"NY",
        "texas":"TX","tx":"TX",
        "california":"CA","ca":"CA","los angeles":"CA","san francisco":"CA",
        "washington":"WA","wa":"WA","seattle":"WA",
        "colorado":"CO","co":"CO","denver":"CO",
        "ohio":"OH","oh":"OH",
        "michigan":"MI","mi":"MI","detroit":"MI",
        "tennessee":"TN","tn":"TN","nashville":"TN",
        "virginia":"VA","va":"VA",
        "maryland":"MD","md":"MD","baltimore":"MD",
        "florida":"FL","fl":"FL","miami":"FL",
        "massachusetts":"MA","ma":"MA","boston":"MA",
        "arizona":"AZ","az":"AZ","phoenix":"AZ",
        "missouri":"MO","mo":"MO","st. louis":"MO","saint louis":"MO",
        "indiana":"IN","in":"IN","indianapolis":"IN",
        "wisconsin":"WI","wi":"WI","milwaukee":"WI",
    }
    # Approximate state centroids for SVG (x,y on 960x600 AlbersUSa-like projection)
    STATE_POS = {
        "AL":(600,360),"AK":(120,480),"AZ":(220,340),"AR":(550,340),
        "CA":(110,270),"CO":(290,280),"CT":(820,210),"DE":(790,250),
        "FL":(660,420),"GA":(640,360),"HI":(240,520),"ID":(210,190),
        "IL":(570,250),"IN":(610,250),"IA":(530,230),"KS":(460,290),
        "KY":(630,290),"LA":(550,390),"ME":(860,160),"MD":(770,260),
        "MA":(840,200),"MI":(620,210),"MN":(510,170),"MS":(580,370),
        "MO":(540,290),"MT":(260,160),"NE":(440,250),"NV":(175,265),
        "NH":(840,185),"NJ":(800,240),"NM":(280,340),"NY":(780,210),
        "NC":(710,310),"ND":(440,165),"OH":(660,240),"OK":(470,330),
        "OR":(145,200),"PA":(740,230),"RI":(845,210),"SC":(680,340),
        "SD":(440,210),"TN":(620,320),"TX":(440,390),"UT":(240,280),
        "VT":(820,185),"VA":(730,280),"WA":(155,160),"WV":(700,265),
        "WI":(560,200),"WY":(295,220),
    }
    import collections as _col
    geo_counts = a.get("geo_counts", {})
    state_counts = _col.defaultdict(int)
    for geo, cnt in geo_counts.items():
        key = str(geo).lower().strip()
        matched = None
        for k, v in STATE_MAP.items():
            if k in key:
                matched = v; break
        if matched:
            state_counts[matched] += cnt
        elif len(key) == 2 and key.upper().isalpha():
            state_counts[key.upper()] += cnt

    map_svg = ""
    if state_counts:
        max_sc = max(state_counts.values())
        svg_parts = ['<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 960 600" style="width:100%;max-width:800px;display:block;margin:0 auto;">']
        svg_parts.append('<rect width="960" height="600" fill="#0f0e2e"/>')
        # Draw all states as circles at their centroid (simplified — no path data needed)
        for state, (cx, cy) in STATE_POS.items():
            count = state_counts.get(state, 0)
            if count > 0:
                intensity = 0.3 + 0.7 * (count / max_sc)
                r = 16 + int(24 * (count / max_sc))
                svg_parts.append('<circle cx="' + str(cx) + '" cy="' + str(cy) + '" r="' + str(r) + '" fill="rgba(40,37,190,' + str(round(intensity,2)) + ')" stroke="#2825BE" stroke-width="1"/>')
                svg_parts.append('<text x="' + str(cx) + '" y="' + str(cy+4) + '" text-anchor="middle" font-family="IBM Plex Mono" font-size="9" fill="white">' + state + '</text>')
            else:
                svg_parts.append('<circle cx="' + str(cx) + '" cy="' + str(cy) + '" r="8" fill="rgba(255,255,255,0.05)" stroke="rgba(255,255,255,0.1)" stroke-width="0.5"/>')
                svg_parts.append('<text x="' + str(cx) + '" y="' + str(cy+3) + '" text-anchor="middle" font-family="IBM Plex Mono" font-size="7" fill="rgba(255,255,255,0.2)">' + state + '</text>')
        # Legend
        svg_parts.append('<text x="20" y="580" font-family="IBM Plex Mono" font-size="9" fill="rgba(255,255,255,0.3)">Circle size and opacity = number of respondents. States with respondents shown in indigo.</text>')
        svg_parts.append('</svg>')
        map_svg = "\n".join(svg_parts)

    # Geo bar fallback (also used as supplementary list)
    geo_html = ""
    if geo_counts:
        max_geo = max(geo_counts.values())
        for g, c in list(geo_counts.items())[:8]:
            pct = min(100, int(c / max_geo * 100))
            geo_html += (
                '<div class="geo-row">'
                '<span class="geo-name">' + str(g) + '</span>'
                '<span class="geo-bar"><span class="geo-fill" style="width:' + str(pct) + '%;"></span></span>'
                '<span class="geo-n">' + str(c) + '</span>'
                '</div>'
            )

    # ── Clusters ───────────────────────────────────────────────────────────
    comm_html = ""
    comm_colors = ["#2825BE", "#0C7A7A", "#EB9001", "#2E9E5B"]
    for i, c in enumerate(communities[:4]):
        size = len(c)
        color = comm_colors[i % len(comm_colors)]
        comm_html += (
            '<div class="comm-card" style="border-top:3px solid ' + color + ';">'
            '<div class="comm-num" style="color:' + color + ';">Cluster ' + str(i+1) + '</div>'
            '<div class="comm-size">' + str(size) + ' member' + ('s' if size != 1 else '') + '</div>'
            '<div class="comm-desc">A group more connected internally than to the rest of the network.</div>'
            '</div>'
        )

    # ── Skill gaps ─────────────────────────────────────────────────────────
    skill_html = ""
    for m in a.get("top_skill_gaps", []):
        need = m["needed"]; off = m["offered"]
        w_need = str(min(100, need * 16)) + "%"
        w_off  = str(min(100, off  * 16)) + "%"
        skill_html += (
            '<div class="skill-row">'
            '<div class="skill-name">' + m["skill"] + '</div>'
            '<div class="skill-bars">'
            '<div class="skill-bar-label">Need (' + str(need) + ')</div>'
            '<div class="skill-bar-wrap"><div class="skill-bar-fill" style="width:' + w_need + ';background:#EB9001;"></div></div>'
            '<div class="skill-bar-label">Offer (' + str(off) + ')</div>'
            '<div class="skill-bar-wrap"><div class="skill-bar-fill" style="width:' + w_off + ';background:#0C7A7A;"></div></div>'
            '</div>'
            '</div>'
        )

    # ── Anonymized network SVG ─────────────────────────────────────────────
    network_svg = ""
    try:
        import networkx as _nx, math as _math
        if len(edge_df) > 0:
            G_html = _nx.DiGraph()
            for _, r in edge_df.iterrows():
                frm = str(r.get("From","")).strip(); to = str(r.get("To","")).strip()
                if frm and to and frm not in ("nan","None") and to not in ("nan","None") and frm != to:
                    G_html.add_edge(frm, to)
            if len(G_html.nodes) > 1:
                UG = G_html.to_undirected()
                pos = _nx.spring_layout(UG, seed=42, k=2.4/_math.sqrt(max(len(G_html.nodes),1)))
                bw  = _nx.betweenness_centrality(G_html)
                deg = dict(G_html.degree())
                indeg = dict(G_html.in_degree())
                max_bw  = max(bw.values())  if bw  else 1
                max_deg = max(deg.values()) if deg else 1
                all_x = [p[0] for p in pos.values()]; all_y = [p[1] for p in pos.values()]
                min_x,max_x = min(all_x),max(all_x); min_y,max_y = min(all_y),max(all_y)
                W,H,PAD = 800,460,48
                def sx(x): return PAD + (x-min_x)/(max_x-min_x+0.001)*(W-2*PAD)
                def sy(y): return PAD + (y-min_y)/(max_y-min_y+0.001)*(H-2*PAD)
                svg_parts = ['<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 ' + str(W) + ' ' + str(H) + '" style="width:100%;background:#0f0e2e;border-radius:4px;">']
                bridge_set = {n for n in G_html.nodes if bw.get(n,0)/max(max_bw,0.001) > 0.12}
                for u,v in G_html.edges():
                    x1,y1=sx(pos[u][0]),sy(pos[u][1]); x2,y2=sx(pos[v][0]),sy(pos[v][1])
                    is_bridge = u in bridge_set or v in bridge_set
                    color = "rgba(235,144,1,0.5)" if is_bridge else "rgba(40,37,190,0.22)"
                    width = "1.8" if is_bridge else "0.9"
                    svg_parts.append('<line x1="' + str(round(x1,1)) + '" y1="' + str(round(y1,1)) + '" x2="' + str(round(x2,1)) + '" y2="' + str(round(y2,1)) + '" stroke="' + color + '" stroke-width="' + width + '"/>')
                for n in G_html.nodes:
                    cx,cy=sx(pos[n][0]),sy(pos[n][1])
                    bw_norm = bw.get(n,0)/max(max_bw,0.001)
                    r = 5 + 14*(deg.get(n,0)/max(max_deg,1))
                    if indeg.get(n,0) == 0: fill = "#CF4C38"
                    elif bw_norm > 0.12:    fill = "#EB9001"
                    else:
                        op = 0.35 + 0.65*(deg.get(n,0)/max(max_deg,1))
                        fill = "rgba(40,37,190," + str(round(op,2)) + ")"
                    svg_parts.append('<circle cx="' + str(round(cx,1)) + '" cy="' + str(round(cy,1)) + '" r="' + str(round(r,1)) + '" fill="' + fill + '" stroke="white" stroke-width="1.2"/>')
                lx = 16
                for lc, ll in [("#EB9001","High bridge role"), ("rgba(40,37,190,0.85)","Connected node"), ("#CF4C38","Not yet named by others")]:
                    svg_parts.append('<circle cx="' + str(lx) + '" cy="' + str(H-14) + '" r="5" fill="' + lc + '"/>')
                    svg_parts.append('<text x="' + str(lx+12) + '" y="' + str(H-10) + '" font-family="IBM Plex Mono" font-size="9" fill="rgba(255,255,255,0.35)">' + ll + '</text>')
                    lx += 170
                svg_parts.append("</svg>")
                network_svg = "\n".join(svg_parts)
    except Exception:
        pass

    # ── Seven-component framing ────────────────────────────────────────────
    comp_rows = [
        ("#2825BE", "2 · Recognition", "Density of weak ties",
         "How many people in this network know each other — and are known by each other. Recognition is the on-ramp to every deeper relational component."),
        ("#0C7A7A", "4 · Mutual Obligation", "Trust under stress · Reciprocity of care",
         "What recognition becomes when activated by trust. The network data measures this directly: trust, energy, support, and reciprocity scores reflect whether mutual obligation is real or only claimed."),
        ("#EB9001", "6 · Institutional Anchors", "Distribution across function",
         "The organizations that hold relational life together. Sector distribution tells us whether coverage spans the full range of functions — or concentrates in a few."),
    ]
    comp_html = ""
    for color, title, attr, desc in comp_rows:
        comp_html += (
            '<div class="comp-card" style="border-left:3px solid ' + color + ';">'
            '<div class="comp-num" style="color:' + color + ';">' + title + '</div>'
            '<div class="comp-attr">' + attr + '</div>'
            '<div class="comp-desc">' + desc + '</div>'
            '</div>'
        )

    # ── Quality cells ──────────────────────────────────────────────────────
    def qcell(label, pct, low_pct, high_msg, low_msg):
        color = "#2E9E5B" if pct >= 60 else ("#EB9001" if pct >= 40 else "#CF4C38")
        sub = high_msg if pct >= 60 else low_msg.replace("{low}", str(low_pct))
        return (
            '<div class="quality-cell">'
            '<div class="qv" style="color:' + color + ';">' + str(pct) + '%</div>'
            '<div class="qinfo"><div class="ql">' + label + '</div>'
            '<div class="qsub">' + sub + '</div></div>'
            '</div>'
        )
    quality_html = (
        qcell("Trust",      pct_trust,      pct_low_trust,
              "Most connections feel trustworthy.",
              "{low}% score low — Cross &amp; Parker identify trust as the dimension most likely to suppress knowledge flow even when formal ties exist.") +
        qcell("Energy",     pct_energy,     pct_low_energy,
              "Connections feel generative and activating.",
              "{low}% feel low-energy. Cross &amp; Parker found de-energizing relationships reduce engagement across the whole network, not just the dyad.") +
        qcell("Support",    pct_support,    0,
              "Strong sense of mutual support across the network.",
              "Support is unevenly distributed — some relationships carry more weight than others.") +
        qcell("Creativity", pct_creativity, 0,
              "Connections feel creatively alive.",
              "Creative spark is lower than other dimensions — worth asking what conditions produce it.")
    )

    # ── Frequency insight ──────────────────────────────────────────────────
    freq_color = "#2E9E5B" if pct_frequent >= 50 else ("#EB9001" if pct_frequent >= 30 else "#CF4C38")
    if pct_frequent >= 60:
        freq_insight = "Most connections are high-frequency — weekly or more. This is a dense relational network in terms of contact, but frequency alone doesn't predict quality. Cross &amp; Parker showed that some of the most frequent contacts are the most draining."
    elif pct_frequent >= 30:
        freq_insight = str(pct_frequent) + "% of connections are weekly or more. The remaining " + str(pct_infrequent) + "% are monthly or less. Infrequent connections aren't necessarily weak — but they are at risk of fading without intentional maintenance."
    else:
        freq_insight = "Most connections are infrequent — monthly or less. Frequency is one of the most reliable predictors of relationship depth over time. The network has dormant capacity that intentional convening could activate."

    # ── Expansion question HTML ────────────────────────────────────────────
    # Collect expansion responses from ego_df if available
    expansion_html = ""
    if "Expansion / Other Members" in ego_df.columns:
        exp_vals = ego_df["Expansion / Other Members"].dropna()
        exp_vals = exp_vals[exp_vals.astype(str).str.strip().str.len() > 2]
        if len(exp_vals) > 0:
            expansion_html = (
                '<div class="expansion-note">'
                '<div class="exp-label">Members named by respondents as missing from the network</div>'
                '<p class="exp-body">Some respondents identified organizations or people they wish TNN had access to. '
                'These responses are shared only in aggregate — no individual attribution — and only where respondents consented to sharing.</p>'
                '<div class="exp-count">' + str(len(exp_vals)) + ' respondent' + ('s' if len(exp_vals) != 1 else '') + ' offered suggestions</div>'
                '</div>'
            )

    # ── Questions for the room ─────────────────────────────────────────────
    questions = []
    if pct_low_trust > 20:
        questions.append(("Trust", str(pct_low_trust) + "% of connections score low on trust.",
            "What would it take to deepen trust in relationships that feel thin? What has worked in your context?"))
    if pct_low_energy > 25:
        questions.append(("Energy", str(pct_low_energy) + "% of connections feel low-energy.",
            "Which relationships feel draining rather than generative? What conditions tend to produce energy in your work?"))
    if pct_balanced < 50:
        questions.append(("Reciprocity", "Many connections flow primarily in one direction.",
            "Where do you notice imbalance in your working relationships? What would more mutual exchange look like?"))
    if n_isolated > 0:
        questions.append(("Isolation", "Some members have not yet been named by others.",
            "Who in this room do you know least well? What has kept you from connecting?"))
    if n_geographies > 2:
        questions.append(("Geography", "This network spans " + str(n_geographies) + " geographic areas.",
            "Where does distance make collaboration harder? What practices help you stay connected across geography?"))
    if n_communities > 1:
        questions.append(("Clusters", "The network has " + str(n_communities) + " distinct clusters.",
            "Which groups in this room interact least with each other? What sits between them?"))
    if pct_deep < 50:
        questions.append(("Depth", "Most connections are still at early stages.",
            "Which relationships here have real potential to deepen? What would cooperation or collaboration actually look like?"))
    if pct_frequent < 40:
        questions.append(("Frequency", "Many connections are infrequent — monthly or less.",
            "What would it take to be in more regular contact? What rhythms — weekly, seasonal, annual — already exist that could carry more?"))
    if not questions:
        questions = [("Ecosystem", "The network is taking shape.",
            "Who in this room do you know least well? What would it mean to strengthen that connection?")]
    # Always add expansion question at end
    questions.append(("Expanding the ecosystem", "What's missing from this network?",
        "Is there an organization, sector, or type of expertise you wish TNN had access to? What would change if they were in the room?"))

    q_html = ""
    for qt, qs, qp in questions:
        tag_color = "#CF4C38" if qt in ["Trust","Isolation","Energy"] else "#EB9001"
        q_html += (
            '<div class="question-card">'
            '<div class="q-tag" style="background:' + tag_color + ';">' + qt + '</div>'
            '<div class="q-stat">' + qs + '</div>'
            '<div class="q-prompt">' + qp + '</div>'
            '</div>'
        )

    # ── HTML ───────────────────────────────────────────────────────────────
    html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>TNN Network Insights</title>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@700;800;900&family=IBM+Plex+Sans:wght@300;400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{--indigo:#2825BE;--amber:#EB9001;--teal:#0C7A7A;--terra:#CF4C38;--green:#2E9E5B;
  --ink:#080818;--ink2:#0f0e2e;--ink3:#13123a;}
*{box-sizing:border-box;margin:0;padding:0;}
body{font-family:'IBM Plex Sans',sans-serif;background:var(--ink);color:white;}
.page{max-width:1120px;margin:0 auto;padding:64px 48px;}
.cover{min-height:100vh;display:flex;flex-direction:column;justify-content:flex-end;padding:96px 48px;border-bottom:1px solid rgba(255,255,255,0.08);}
.cover-eyebrow{font-family:'IBM Plex Mono',monospace;font-size:11px;letter-spacing:0.18em;text-transform:uppercase;color:var(--amber);margin-bottom:24px;}
.cover-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(52px,7vw,96px);font-weight:900;line-height:0.95;letter-spacing:-0.02em;text-transform:uppercase;color:white;margin-bottom:24px;}
.cover-sub{font-family:'IBM Plex Sans',sans-serif;font-size:18px;color:rgba(255,255,255,0.5);max-width:520px;line-height:1.6;}
.cover-meta{font-family:'IBM Plex Mono',monospace;font-size:11px;color:rgba(255,255,255,0.3);margin-top:48px;letter-spacing:0.1em;}
.section{padding:80px 0;border-bottom:1px solid rgba(255,255,255,0.08);}
.section:last-child{border-bottom:none;}
.section-eyebrow{font-family:'IBM Plex Mono',monospace;font-size:10px;letter-spacing:0.18em;text-transform:uppercase;color:var(--amber);margin-bottom:12px;}
.section-title{font-family:'Barlow Condensed',sans-serif;font-size:clamp(32px,4vw,52px);font-weight:800;text-transform:uppercase;letter-spacing:-0.01em;color:white;margin-bottom:32px;line-height:1.1;}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:40px;align-items:start;}
.stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:2px;margin:32px 0;}
.stat-cell{background:var(--ink2);padding:32px 24px;}
.stat-cell .sv{font-family:'Barlow Condensed',sans-serif;font-size:56px;font-weight:900;line-height:1;color:var(--indigo);}
.stat-cell .sl{font-family:'IBM Plex Mono',monospace;font-size:10px;color:rgba(255,255,255,0.3);text-transform:uppercase;letter-spacing:0.1em;margin-top:6px;}
.quality-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:2px;margin:24px 0;}
.quality-cell{background:var(--ink2);padding:24px;display:flex;align-items:center;gap:20px;}
.quality-cell .qv{font-family:'Barlow Condensed',sans-serif;font-size:44px;font-weight:800;min-width:80px;line-height:1;}
.quality-cell .qinfo .ql{font-family:'IBM Plex Mono',monospace;font-size:10px;color:rgba(255,255,255,0.3);text-transform:uppercase;letter-spacing:0.1em;}
.quality-cell .qinfo .qsub{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.55);margin-top:4px;line-height:1.5;}
.kesc-note{background:var(--ink3);border-left:3px solid var(--teal);padding:20px 24px;margin:24px 0;}
.kesc-note .kn-label{font-family:'IBM Plex Mono',monospace;font-size:9px;color:var(--teal);letter-spacing:0.14em;text-transform:uppercase;margin-bottom:6px;}
.kesc-note .kn-text{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.5);line-height:1.65;}
.freq-cell{background:var(--ink2);padding:28px 32px;margin:24px 0;display:flex;align-items:center;gap:32px;}
.freq-cell .fv{font-family:'Barlow Condensed',sans-serif;font-size:52px;font-weight:800;line-height:1;min-width:100px;}
.freq-cell .fi{font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:rgba(255,255,255,0.55);line-height:1.65;max-width:520px;}
.depth-bar{background:var(--ink2);padding:32px;margin:16px 0;}
.depth-bar-label{font-family:'IBM Plex Mono',monospace;font-size:10px;color:rgba(255,255,255,0.3);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:16px;}
.depth-segments{display:flex;height:40px;overflow:hidden;gap:2px;}
.depth-seg{display:flex;align-items:center;justify-content:center;font-family:'IBM Plex Mono',monospace;font-size:10px;color:white;}
.depth-legend{display:flex;gap:24px;margin-top:12px;flex-wrap:wrap;}
.dl-item{display:flex;align-items:center;gap:6px;font-family:'IBM Plex Sans',sans-serif;font-size:12px;color:rgba(255,255,255,0.4);}
.dl-dot{width:8px;height:8px;border-radius:50%;}
.comp-grid{display:flex;flex-direction:column;gap:2px;margin:24px 0;}
.comp-card{background:var(--ink2);padding:24px;}
.comp-num{font-family:'IBM Plex Mono',monospace;font-size:10px;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;}
.comp-attr{font-family:'Barlow Condensed',sans-serif;font-size:18px;font-weight:700;color:white;margin-bottom:6px;}
.comp-desc{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.45);line-height:1.6;}
.chip-row{display:flex;flex-wrap:wrap;gap:8px;margin:16px 0;}
.chip{background:var(--ink2);border:1px solid rgba(255,255,255,0.1);padding:6px 12px;font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.7);display:flex;align-items:center;gap:8px;}
.chip-n{font-family:'IBM Plex Mono',monospace;font-size:11px;color:var(--amber);}
.geo-row{display:flex;align-items:center;gap:12px;margin:8px 0;}
.geo-name{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.6);min-width:160px;}
.geo-bar{flex:1;height:5px;background:rgba(255,255,255,0.07);overflow:hidden;}
.geo-fill{height:100%;background:var(--indigo);}
.geo-n{font-family:'IBM Plex Mono',monospace;font-size:11px;color:rgba(255,255,255,0.3);min-width:20px;text-align:right;}
.comm-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:2px;margin:24px 0;}
.comm-card{background:var(--ink2);padding:24px;}
.comm-num{font-family:'IBM Plex Mono',monospace;font-size:10px;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px;}
.comm-size{font-family:'Barlow Condensed',sans-serif;font-size:32px;font-weight:800;color:white;line-height:1;}
.comm-desc{font-family:'IBM Plex Sans',sans-serif;font-size:12px;color:rgba(255,255,255,0.3);margin-top:6px;line-height:1.5;}
.skill-row{margin:20px 0;}
.skill-name{font-family:'IBM Plex Sans',sans-serif;font-size:15px;color:white;margin-bottom:8px;}
.skill-bar-label{font-family:'IBM Plex Mono',monospace;font-size:9px;color:rgba(255,255,255,0.3);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:3px;}
.skill-bars{margin-bottom:4px;}
.skill-bar-wrap{height:8px;background:rgba(255,255,255,0.06);margin-bottom:3px;width:100%;}
.skill-bar-fill{height:100%;}
.network-label{font-family:'IBM Plex Mono',monospace;font-size:10px;color:rgba(255,255,255,0.25);text-transform:uppercase;letter-spacing:0.12em;margin-bottom:10px;}
.expansion-note{background:var(--ink3);border:1px solid rgba(235,144,1,0.3);padding:24px;margin:24px 0;}
.exp-label{font-family:'IBM Plex Mono',monospace;font-size:10px;color:var(--amber);text-transform:uppercase;letter-spacing:0.12em;margin-bottom:8px;}
.exp-body{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.5);line-height:1.65;margin-bottom:12px;}
.exp-count{font-family:'Barlow Condensed',sans-serif;font-size:28px;font-weight:800;color:var(--amber);}
.question-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:2px;margin:24px 0;}
.question-card{background:var(--ink2);padding:28px;}
.q-tag{display:inline-block;font-family:'IBM Plex Mono',monospace;font-size:9px;letter-spacing:0.14em;text-transform:uppercase;color:white;padding:3px 8px;margin-bottom:14px;}
.q-stat{font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.4);margin-bottom:10px;line-height:1.5;}
.q-prompt{font-family:'IBM Plex Sans',sans-serif;font-size:15px;color:white;line-height:1.65;font-style:italic;}
.closing{background:var(--ink2);padding:64px 48px;border-top:3px solid var(--amber);}
.closing-title{font-family:'Barlow Condensed',sans-serif;font-size:44px;font-weight:800;text-transform:uppercase;color:white;margin-bottom:16px;line-height:1.1;}
.closing-body{font-family:'IBM Plex Sans',sans-serif;font-size:16px;color:rgba(255,255,255,0.55);max-width:600px;line-height:1.7;}
.footer{font-family:'IBM Plex Mono',monospace;font-size:10px;color:rgba(255,255,255,0.15);text-align:center;padding:32px;letter-spacing:0.1em;}
@media print{body{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
</style>
</head>
<body>

<div class="cover">
  <div class="cover-eyebrow">Connecting for Change · Thriving Neighborhoods Network</div>
  <div class="cover-title">What the<br>Network<br>Shows</div>
  <div class="cover-sub">An anonymized portrait of connections, strengths, and opportunities across the TNN ecosystem — prepared to support Activity 5: Who Is in the Ecosystem?</div>
  <div class="cover-meta">n=""" + str(n_resp) + """ respondents · """ + str(n_nodes) + """ people in the network · """ + str(n_edges) + """ connections · density """ + str(density) + """</div>
</div>

<div class="page">

<!-- SCALE -->
<div class="section">
  <div class="section-eyebrow">Network Scale</div>
  <div class="section-title">The shape<br>of what exists</div>
  <div class="stat-grid">
    <div class="stat-cell"><div class="sv">""" + str(n_resp) + """</div><div class="sl">Respondents</div></div>
    <div class="stat-cell"><div class="sv">""" + str(n_nodes) + """</div><div class="sl">People in the network</div></div>
    <div class="stat-cell"><div class="sv">""" + str(n_edges) + """</div><div class="sl">Named connections</div></div>
    <div class="stat-cell"><div class="sv">""" + str(n_communities) + """</div><div class="sl">Distinct clusters</div></div>
  </div>
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:15px;color:rgba(255,255,255,0.45);max-width:600px;line-height:1.7;">
    Each respondent named people they work with and rated the quality of those relationships.
    This view is aggregated, anonymized, and combined — no individual data is shown.
  </p>
</div>

""" + ("""
<!-- NETWORK MAP -->
<div class="section">
  <div class="section-eyebrow">Network Structure · Anonymized</div>
  <div class="section-title">Who is connected<br>to whom</div>
  <div class="network-label">Node size = connections &nbsp;·&nbsp; Amber = bridge role &nbsp;·&nbsp; Red = not yet named by others &nbsp;·&nbsp; No names shown</div>
  """ + network_svg + """
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:rgba(255,255,255,0.3);max-width:640px;line-height:1.7;margin-top:20px;">
    The amber nodes carry the most connective load — they sit on the most paths between others.
    These are the people Activity 5 asks about: the trusted connectors, the bridge actors, those who translate across worlds.
    The structure matters more than the names.
  </p>
</div>
""" if network_svg else "") + """

<!-- SEVEN COMPONENTS -->
<div class="section">
  <div class="section-eyebrow">Mapped to the seven components</div>
  <div class="section-title">What the network<br>data can say</div>
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:15px;color:rgba(255,255,255,0.45);max-width:640px;line-height:1.7;margin-bottom:32px;">
    The TNN framework identifies seven components that make a neighborhood thrive.
    The network survey speaks most directly to three.
  </p>
  <div class="comp-grid">""" + comp_html + """</div>
</div>

<!-- DEPTH -->
<div class="section">
  <div class="section-eyebrow">Depth of Connection · Mutual Obligation</div>
  <div class="section-title">""" + str(pct_deep) + """% of connections<br>reach cooperation<br>or collaboration</div>
  <div class="depth-bar">
    <div class="depth-bar-label">Distribution across all reported connections</div>
    <div class="depth-segments">
      """ + dc_awareness_html + dc_connection_html + dc_cooperation_html + dc_collaboration_html + """
    </div>
    <div class="depth-legend">
      <div class="dl-item"><div class="dl-dot" style="background:#13123a;"></div>Awareness</div>
      <div class="dl-item"><div class="dl-dot" style="background:#1e1c4a;"></div>Connection</div>
      <div class="dl-item"><div class="dl-dot" style="background:#2825BE;"></div>Cooperation</div>
      <div class="dl-item"><div class="dl-dot" style="background:#4a47d6;"></div>Collaboration</div>
    </div>
  </div>
</div>

<!-- RELATIONAL QUALITY -->
<div class="section">
  <div class="section-eyebrow">Relational Quality · KESC Framework</div>
  <div class="section-title">% rated "quite a bit"<br>or "a great deal"</div>
  <div class="quality-grid">""" + quality_html + """</div>
  <div class="kesc-note">
    <div class="kn-label">Source · Cross &amp; Parker (2004)</div>
    <div class="kn-text">
      In <em>The Hidden Power of Social Networks</em>, Cross and Parker found that what makes a network valuable is not the number of connections, but the quality of energy flowing through them.
      They identified four dimensions — Knowledge, Energy, Safety (trust), and Credibility (support) — that predict whether a relationship generates real value or quietly drains it.
      A frequent contact who scores low on energy and trust is not a network asset; they are a structural liability.
      The KESC lens shifts attention from <em>who knows whom</em> to <em>what flows between them</em>.
    </div>
  </div>
  <p style="font-family:'IBM Plex Mono',monospace;font-size:11px;color:rgba(255,255,255,0.25);margin-top:12px;">
    Reciprocity: """ + str(pct_balanced) + """% described as roughly balanced &nbsp;·&nbsp; """ + str(pct_frequent) + """% interact weekly or more
  </p>
</div>

<!-- FREQUENCY -->
<div class="section">
  <div class="section-eyebrow">Frequency of Interaction</div>
  <div class="section-title">How often<br>connections happen</div>
  <div class="freq-cell">
    <div class="fv" style="color:""" + freq_color + """;">""" + str(pct_frequent) + """%</div>
    <div class="fi">""" + freq_insight + """</div>
  </div>
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:rgba(255,255,255,0.35);max-width:600px;line-height:1.7;margin-top:8px;">
    """ + str(pct_infrequent) + """% of connections are monthly or less. These are not lost relationships — they are dormant ones.
    The question is whether the network has the rhythms and structures to keep them alive.
    Rhythm &amp; Recurrence is Component 3 in the TNN framework for exactly this reason.
  </p>
</div>

<!-- ECOSYSTEM COMPOSITION -->
<div class="section">
  <div class="section-eyebrow">Ecosystem Composition · Activity 5</div>
  <div class="section-title">Who is visible<br>in this network</div>
  <div class="two-col">
    <div>
      """ + ("""<div class="section-eyebrow" style="color:#0C7A7A;margin-bottom:8px;">Sectors represented</div>
      <div class="chip-row">""" + sector_html + """</div>""" if sector_html else "") + """
      """ + ("""<div class="section-eyebrow" style="color:#0C7A7A;margin-top:28px;margin-bottom:8px;">Geographic spread · """ + str(n_geographies) + """ area""" + ("s" if n_geographies != 1 else "") + """</div>""" + geo_html if geo_html else "") + """
    </div>
    <div>
      """ + ("""<div class="section-eyebrow" style="color:#0C7A7A;margin-bottom:8px;">Network clusters</div>
      <div class="comm-grid">""" + comm_html + """</div>
      <p style="font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:rgba(255,255,255,0.3);line-height:1.6;margin-top:12px;">
        Clusters are groups more connected internally than to the rest of the network.
        Activity 5 asks: what sits between them?
      </p>""" if comm_html else "") + """
    </div>
  </div>
  """ + ("""<div class="section-eyebrow" style="color:#0C7A7A;margin-top:32px;margin-bottom:8px;">Geographic distribution</div>""" + map_svg if map_svg else "") + """
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:rgba(255,255,255,0.3);margin-top:24px;max-width:600px;line-height:1.7;">
    The dominant relationship type is <strong style="color:rgba(255,255,255,0.6);">""" + top_rel + """</strong>.
    """ + ("The network spans " + str(n_geographies) + " geographic areas — distance shapes who stays connected." if n_geographies > 2 else "") + """
  </p>
</div>

""" + ("""
<!-- GAPS -->
<div class="section">
  <div class="section-eyebrow" style="color:#CF4C38;">Gaps &amp; Opportunities</div>
  <div class="section-title">Where the network<br>could be stronger</div>
  """ + ("""<div style="margin:24px 0;">
    <div class="section-eyebrow" style="color:#EB9001;margin-bottom:12px;">Skill gaps — what people need that others could offer</div>
    """ + skill_html + """
  </div>""" if skill_html else "") + """
  """ + ("""<div style="display:inline-block;background:#CF4C38;padding:4px 12px;font-family:'IBM Plex Mono',monospace;font-size:11px;color:white;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:12px;">Potential isolation</div>
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:14px;color:rgba(255,255,255,0.5);max-width:580px;line-height:1.7;">""" + str(n_isolated + n_periphery) + """ people appear in only one person's named connections. They may be underconnected relative to their potential role.</p>""" if (n_isolated + n_periphery) > 0 else "") + """
  """ + expansion_html + """
</div>
""" if (skill_html or (n_isolated + n_periphery) > 0 or expansion_html) else "") + """

<!-- QUESTIONS -->
<div class="section">
  <div class="section-eyebrow">For the room · Activity 5</div>
  <div class="section-title">What the data<br>can't answer alone</div>
  <p style="font-family:'IBM Plex Sans',sans-serif;font-size:15px;color:rgba(255,255,255,0.45);max-width:580px;line-height:1.7;margin-bottom:32px;">
    These questions emerge from the network data. They are not conclusions — they are starting points.
    The ecosystem becomes visible when people in the room name what the data cannot.
  </p>
  <div class="question-grid">""" + q_html + """</div>
</div>

</div>

<div class="closing">
  <div class="closing-title">The map is incomplete<br>by design.</div>
  <div class="closing-body">
    What you see here represents what people chose to name. The relationships that matter most,
    the actors who are missing, the conditions that are fragile — those only become visible
    when people in this room say what the data cannot.<br><br>
    Your role is not to validate the map. Your role is to improve it.
  </div>
</div>

<div class="footer">TNN NETWORK INSIGHTS · CONNECTING FOR CHANGE · ANONYMIZED · NOT FOR DISTRIBUTION</div>
</body>
</html>"""
    return html


# ── Skill gap chart (single clean one) ───────────────────────────────────────
def skill_gap_chart(matches):
    skills=[m["skill"] for m in matches[:10]]
    offered=[m["offered"] for m in matches[:10]]
    needed=[m["needed"] for m in matches[:10]]
    fig=go.Figure()
    fig.add_trace(go.Bar(name="Needed",y=skills,x=needed,orientation="h",marker_color=AMBER,marker_line_width=0))
    fig.add_trace(go.Bar(name="Offered",y=skills,x=offered,orientation="h",marker_color=TEAL,marker_line_width=0,opacity=0.8))
    fig.update_layout(barmode="overlay",paper_bgcolor=SURFACE,plot_bgcolor=SURFACE,
        title=dict(text="Skills: Supply vs. Demand",font=dict(family="Barlow Condensed",size=18,color=INK)),
        xaxis=dict(tickfont=dict(family="IBM Plex Mono",size=10),gridcolor=BORDER,linecolor=BORDER),
        yaxis=dict(tickfont=dict(family="IBM Plex Sans",size=11),showgrid=False,autorange="reversed"),
        legend=dict(font=dict(family="IBM Plex Sans",size=11),orientation="h",y=1.08),
        margin=dict(l=16,r=16,t=48,b=16),height=max(300,32*len(skills)+80))
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="eyebrow">Connecting for Change</div>',unsafe_allow_html=True)
st.markdown('<h1 style="font-size:38px;margin:0 0 4px 0;font-family:Barlow Condensed,sans-serif;">TNN Network Analyzer</h1>',unsafe_allow_html=True)
st.markdown(f'<p style="color:{TEXT2};font-size:14px;margin:0 0 20px 0;">Upload participant exports · Explore the network · Export insights</p>',unsafe_allow_html=True)
st.markdown(f'<hr style="border:none;border-top:1px solid {BORDER};margin:0 0 24px 0">',unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="eyebrow">Data Input</div>',unsafe_allow_html=True)
    st.markdown("**NetCanvas exports**")
    uploaded=st.file_uploader("Zip files",accept_multiple_files=True,type=["zip"],label_visibility="collapsed")
    st.markdown("---")
    st.markdown("**Pre-survey** *(optional)*")
    st.markdown(f'<p style="font-size:12px;color:{TEXT2};">Network Activation Survey Excel — adds skills, challenges, and context.</p>',unsafe_allow_html=True)
    presurvey_file=st.file_uploader("Excel",type=["xlsx"],label_visibility="collapsed",key="presurvey")

    if uploaded:
        if st.button("▶  Process",type="primary",use_container_width=True):
            with st.spinner("Processing..."):
                buckets=extract_zips(uploaded)
                ego_df,edge_df,warnings=process_buckets(buckets)
                pre_df=parse_presurvey(presurvey_file) if presurvey_file else pd.DataFrame()
                if not pre_df.empty: ego_df=merge_presurvey(ego_df,pre_df)
                G=build_graph(edge_df)
                ego_names=list(ego_df["Name"].dropna()) if "Name" in ego_df.columns else []
                node_meta={}
                if not pre_df.empty:
                    for _,r in pre_df.iterrows():
                        node_meta[r["name"]]={"challenge":r.get("challenge",""),"skills_offered":r.get("skills_offered","")}
                analytics=compute_analytics(ego_df,edge_df,G,pre_df if not pre_df.empty else None)
                st.session_state.update({"ego_df":ego_df,"edge_df":edge_df,"G":G,
                    "ego_names":ego_names,"pre_df":pre_df,"node_meta":node_meta,"analytics":analytics})
            for w in warnings: st.warning(w)
            st.success(f"✅ {len(ego_df)} participant(s) · {len(edge_df)} edge(s)")

    if "edge_df" in st.session_state and not st.session_state["edge_df"].empty:
        st.markdown("---")
        st.markdown('<div class="eyebrow">Exports</div>',unsafe_allow_html=True)
        excel=build_excel(st.session_state["ego_df"],st.session_state["edge_df"])
        st.download_button("⬇  Excel",data=excel,file_name="tnn_combined_network.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        if "analytics" in st.session_state:
            html_bytes=build_insights_html(st.session_state["analytics"],
                st.session_state["ego_df"],st.session_state["edge_df"]).encode("utf-8")
            st.download_button("⬇  Insights HTML",data=html_bytes,file_name="tnn_network_insights.html",
                mime="text/html",use_container_width=True)

# ── Gate ──────────────────────────────────────────────────────────────────────
if "ego_df" not in st.session_state:
    st.markdown(f"""<div style="background:{SURFACE2};border-radius:8px;padding:48px 32px;text-align:center;border:1px dashed {BORDER};">
        <div style="font-family:'Barlow Condensed',sans-serif;font-size:22px;color:{TEXT2};font-weight:700;text-transform:uppercase;">Upload zip files in the sidebar to get started</div>
        <div style="font-family:'IBM Plex Sans',sans-serif;font-size:13px;color:{TEXT3};margin-top:8px;">Pre-survey Excel is optional — adds skills, challenges, and neighborhood context</div>
    </div>""",unsafe_allow_html=True)
    st.stop()

ego_df=st.session_state["ego_df"]; edge_df=st.session_state["edge_df"]
G=st.session_state["G"]; ego_names=st.session_state.get("ego_names",[])
pre_df=st.session_state.get("pre_df",pd.DataFrame())
node_meta=st.session_state.get("node_meta",{})
a=st.session_state.get("analytics",{})

# ── Metric strip ──────────────────────────────────────────────────────────────
c1,c2,c3,c4=st.columns(4)
for col,(val,lbl_,color) in zip([c1,c2,c3,c4],[
    (len(ego_df),"SURVEY TAKERS",INDIGO),(len(G.nodes),"UNIQUE NODES",TEAL),
    (len(G.edges),"CONNECTIONS",AMBER),(f"{nx.density(G):.3f}","NETWORK DENSITY",GREEN)]):
    with col:
        st.markdown(f'<div class="metric-card" style="border-left-color:{color};"><div class="value" style="color:{color};">{val}</div><div class="label">{lbl_}</div></div>',unsafe_allow_html=True)

st.markdown("<br>",unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1,tab2,tab3,tab4,tab5=st.tabs(["🕸  Network","📊  Connections","🛠  Skills","💡  Insights","🖥  Room View"])

# ═══ TAB 1: NETWORK ══════════════════════════════════════════════════════════
with tab1:
    ctrl_l,ctrl_r=st.columns([4,1])
    with ctrl_r:
        st.markdown("<br>",unsafe_allow_html=True)
        color_by=st.selectbox("Color nodes by",["degree","in_degree","betweenness","survey_taker","isolated"],
            format_func=lambda x:{"degree":"Connections (degree)","in_degree":"Times named","betweenness":"Bridge role","survey_taker":"Respondent vs. named","isolated":"Potential isolation"}.get(x,x))
        show_labels=st.checkbox("Show labels",value=True)
    with ctrl_l:
        st.plotly_chart(make_network_fig(G,color_by=color_by,ego_names=ego_names,node_meta=node_meta,show_labels=show_labels),use_container_width=True)

    col_a,col_b=st.columns(2)
    with col_a:
        st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Centrality</div>',unsafe_allow_html=True)
        deg=dict(G.degree()); indeg=dict(G.in_degree()); bw=nx.betweenness_centrality(G)
        cent=pd.DataFrame([{"Name":n,"Degree":deg[n],"Times Named":indeg[n],"Betweenness":round(bw[n],3)} for n in G.nodes])
        st.dataframe(cent.sort_values("Degree",ascending=False).reset_index(drop=True),use_container_width=True,hide_index=True)
    with col_b:
        st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Clusters · Louvain</div>',unsafe_allow_html=True)
        UG2=G.to_undirected()
        if len(UG2.nodes)>1:
            try:
                UG2_clean=nx.Graph()
                for u,v in UG2.edges():
                    if isinstance(u,str) and isinstance(v,str) and u!=v:
                        UG2_clean.add_edge(u,v)
                for n in UG2.nodes():
                    if isinstance(n,str): UG2_clean.add_node(n)
                comms=nx.community.louvain_communities(UG2_clean,seed=42)
                rows=[{"Cluster":f"Cluster {i+1}","Size":len(c),"Members":", ".join(sorted(list(c)))} for i,c in enumerate(sorted(comms,key=len,reverse=True))]
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
            except: st.info("Community detection unavailable.")

# ═══ TAB 2: CONNECTIONS ══════════════════════════════════════════════════════
with tab2:
    if edge_df.empty: st.info("No edge data.")
    else:
        # Scorecards for relational quality
        attrs=["Trust","Energy","Support","Creativity"]
        scores=[a.get(f"pct_{x.lower()}",0) for x in attrs]
        colors=[GREEN if s>=60 else (AMBER if s>=40 else TERRA) for s in scores]
        sub_labels=[]
        for attr,s in zip(attrs,scores):
            if attr=="Trust": sub_labels.append("high-trust" if s>=60 else f"{a.get('pct_low_trust',0)}% score low")
            elif attr=="Energy": sub_labels.append("feel generative" if s>=60 else f"{a.get('pct_low_energy',0)}% feel low-energy")
            else: sub_labels.append("" )

        sc_html='<div class="scorecard-row">'
        sc_html = '<div class="scorecard-row">'
        for attr,s,c,sub in zip(attrs,scores,colors,sub_labels):
            sub_div = f'<div class="sc-sub">{sub}</div>' if sub else ''
            sc_html += f'<div class="scorecard" style="border-top-color:{c};"><div class="sc-val" style="color:{c};">{s}%</div><div class="sc-label">{attr}</div>{sub_div}</div>'
        sc_html += '</div>'


        st.markdown(f'<div class="eyebrow">Relational Quality — % scoring "Quite a bit" or "A great deal"</div>',unsafe_allow_html=True)
        st.markdown(sc_html,unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        col_a,col_b=st.columns(2)

        with col_a:
            st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Depth of connection</div>',unsafe_allow_html=True)
            if "Depth of Connection" in edge_df.columns:
                order=["Awareness","Connection","Cooperation","Collaboration"]
                counts=edge_df["Depth of Connection"].value_counts().reindex([o for o in order if o in edge_df["Depth of Connection"].values],fill_value=0)
                fig=go.Figure(go.Bar(x=list(counts.index),y=list(counts.values),
                    marker_color=[SURFACE2,"#c7c5f5",INDIGO_LT,INDIGO],marker_line_width=0,
                    text=list(counts.values),textposition="outside",textfont=dict(family="IBM Plex Mono",size=10)))
                fig.update_layout(paper_bgcolor=SURFACE,plot_bgcolor=SURFACE,showlegend=False,
                    xaxis=dict(tickfont=dict(family="IBM Plex Mono",size=10,color=TEXT2),showgrid=False),
                    yaxis=dict(showgrid=False,showticklabels=False),
                    margin=dict(l=0,r=0,t=16,b=0),height=200)
                st.plotly_chart(fig,use_container_width=True)

            st.markdown(f'<div class="eyebrow" style="color:{TEAL};margin-top:16px;">Reciprocity</div>',unsafe_allow_html=True)
            if "Reciprocity" in edge_df.columns:
                rc=edge_df["Reciprocity"].value_counts()
                fig=go.Figure(go.Bar(x=list(rc.values),y=list(rc.index),orientation="h",
                    marker_color=[GREEN if "balanced" in str(k).lower() else AMBER for k in rc.index],
                    marker_line_width=0,text=list(rc.values),textposition="outside",
                    textfont=dict(family="IBM Plex Mono",size=10)))
                fig.update_layout(paper_bgcolor=SURFACE,plot_bgcolor=SURFACE,showlegend=False,
                    xaxis=dict(showgrid=False,showticklabels=False),
                    yaxis=dict(tickfont=dict(family="IBM Plex Sans",size=11,color=TEXT2),showgrid=False,autorange="reversed"),
                    margin=dict(l=0,r=0,t=8,b=0),height=200)
                st.plotly_chart(fig,use_container_width=True)

        with col_b:
            st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Relationship types</div>',unsafe_allow_html=True)
            if "Relationship Type(s)" in edge_df.columns:
                all_t=[]
                for v in edge_df["Relationship Type(s)"].dropna():
                    all_t.extend([t.strip() for t in str(v).split(";") if t.strip()])
                if all_t:
                    tc=pd.Series(all_t).value_counts()
                    short_labels=[l.split("/")[0].strip() for l in tc.index]
                    fig=go.Figure(go.Bar(x=list(tc.values),y=short_labels,orientation="h",
                        marker_color=TEAL,marker_line_width=0,
                        text=list(tc.values),textposition="outside",textfont=dict(family="IBM Plex Mono",size=10)))
                    fig.update_layout(paper_bgcolor=SURFACE,plot_bgcolor=SURFACE,showlegend=False,
                        xaxis=dict(showgrid=False,showticklabels=False),
                        yaxis=dict(tickfont=dict(family="IBM Plex Sans",size=11,color=TEXT2),showgrid=False,autorange="reversed"),
                        margin=dict(l=0,r=0,t=8,b=0),height=260)
                    st.plotly_chart(fig,use_container_width=True)

            st.markdown(f'<div class="eyebrow" style="color:{TEAL};margin-top:16px;">Frequency</div>',unsafe_allow_html=True)
            if "Frequency of Interaction" in edge_df.columns:
                order=["Never or rarely","Once or twice a month","Weekly","Multiple times a week","Daily or near-daily"]
                fc=edge_df["Frequency of Interaction"].value_counts().reindex([o for o in order if o in edge_df["Frequency of Interaction"].values],fill_value=0)
                fig=go.Figure(go.Bar(x=[o.split(" ")[0] for o in fc.index],y=list(fc.values),
                    marker_color=TEAL,marker_line_width=0,
                    text=list(fc.values),textposition="outside",textfont=dict(family="IBM Plex Mono",size=10)))
                fig.update_layout(paper_bgcolor=SURFACE,plot_bgcolor=SURFACE,showlegend=False,
                    xaxis=dict(tickfont=dict(family="IBM Plex Mono",size=10,color=TEXT2),showgrid=False),
                    yaxis=dict(showgrid=False,showticklabels=False),
                    margin=dict(l=0,r=0,t=8,b=0),height=180)
                st.plotly_chart(fig,use_container_width=True)

# ═══ TAB 3: SKILLS ═══════════════════════════════════════════════════════════
with tab3:
    if pre_df.empty:
        st.info("Upload the pre-survey Excel in the sidebar to enable this tab.")
    else:
        matches=build_skill_match(pre_df)
        st.plotly_chart(skill_gap_chart(matches),use_container_width=True)

        matchable=[m for m in matches if m["offered"]>0 and m["needed"]>0]
        if matchable:
            st.markdown(f'<div class="eyebrow" style="color:{TEAL};margin-top:8px;">Skill matchmaking</div>',unsafe_allow_html=True)
            st.markdown(f'<p style="font-size:13px;color:{TEXT2};margin-bottom:16px;">Lines connect people who offer a skill to those who need it. The thicker the opportunity, the more people it could connect.</p>',unsafe_allow_html=True)

            # Build SVG matchmaking diagram using helper function
            st.markdown(build_skill_svg(matchable[:6]), unsafe_allow_html=True)

        if "challenge" in pre_df.columns:
            st.markdown(f'<div class="eyebrow" style="color:{TERRA};margin-top:16px;">Challenges named</div>',unsafe_allow_html=True)
            ch=pre_df[["name","organization","challenge"]].dropna(subset=["challenge"]).rename(columns={"name":"Name","organization":"Organization","challenge":"Challenge"})
            st.dataframe(ch,use_container_width=True,hide_index=True)

# ═══ TAB 4: INSIGHTS ═════════════════════════════════════════════════════════
with tab4:
    if not a:
        st.info("Process data to generate insights.")
    else:
        # ── Hub-style header ──
        st.markdown(f"""
        <div style="background:{INK};border-radius:8px;padding:40px 40px 32px;margin-bottom:24px;">
          <div class="eyebrow" style="color:{AMBER};">Network Insights · TNN 2026</div>
          <h2 style="font-family:Barlow Condensed,sans-serif;font-size:44px;font-weight:900;text-transform:uppercase;color:white;letter-spacing:-0.02em;margin:8px 0 12px;line-height:1;">What the network shows</h2>
          <p style="font-family:IBM Plex Sans,sans-serif;font-size:15px;color:rgba(255,255,255,0.5);max-width:560px;line-height:1.7;margin:0;">
            {a.get("n_respondents",0)} respondents · {a.get("n_nodes",0)} people in the network · {a.get("n_edges",0)} connections · density {a.get("density",0):.3f}
          </p>
        </div>""", unsafe_allow_html=True)

        # ── Seven-component lens ──
        st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Mapped to the seven components framework</div>', unsafe_allow_html=True)
        st.markdown(f'<p style="font-size:13px;color:{TEXT2};max-width:620px;line-height:1.6;margin-bottom:16px;">The TNN framework identifies seven components that make a neighborhood thrive. The network data speaks most directly to the relational components: Recognition, Mutual Obligation, and Institutional Anchors.</p>', unsafe_allow_html=True)

        trust_pct = a.get("pct_trust",0)
        trust_low = a.get("pct_low_trust",0)
        energy_pct = a.get("pct_energy",0)
        energy_low = a.get("pct_low_energy",0)
        pct_deep = a.get("pct_deep",0)
        pct_balanced = a.get("pct_balanced",0)
        n_communities = a.get("n_communities",1)
        n_geographies = a.get("n_geographies",0)

        COMP_COLORS = {"recognition":INDIGO,"mutual_obligation":TEAL,"institutional":AMBER,"common_stakes":GREEN,"geometry":TERRA}

        def insight_card(eyebrow, eyebrow_color, headline, stat, stat_color, body, component_tag=None):
            tag_html = f'<div style="font-family:IBM Plex Mono,monospace;font-size:9px;letter-spacing:0.12em;text-transform:uppercase;color:{eyebrow_color};border:1px solid {eyebrow_color};padding:2px 8px;border-radius:2px;display:inline-block;margin-bottom:8px;">{component_tag}</div>' if component_tag else ""
            return f"""<div style="border-left:3px solid {eyebrow_color};padding:20px 24px;background:{SURFACE2};margin-bottom:12px;border-radius:0 6px 6px 0;">
              {tag_html}
              <div style="font-family:IBM Plex Mono,monospace;font-size:9px;letter-spacing:0.14em;text-transform:uppercase;color:{eyebrow_color};margin-bottom:6px;">{eyebrow}</div>
              <div style="font-family:Barlow Condensed,sans-serif;font-size:28px;font-weight:800;color:{stat_color};line-height:1;margin-bottom:6px;">{stat}</div>
              <div style="font-family:Barlow Condensed,sans-serif;font-size:18px;font-weight:700;color:{INK};margin-bottom:6px;">{headline}</div>
              <div style="font-family:IBM Plex Sans,sans-serif;font-size:13px;color:{TEXT2};line-height:1.65;max-width:560px;">{body}</div>
            </div>"""

        col_l, col_r = st.columns(2)

        with col_l:
            # Recognition component
            st.markdown(f'<div style="font-family:Barlow Condensed,sans-serif;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:{INDIGO};border-bottom:2px solid {INDIGO};padding-bottom:6px;margin-bottom:12px;">2 · Recognition</div>', unsafe_allow_html=True)
            indeg = dict(G.in_degree())
            n_named_multi = sum(1 for n,d in indeg.items() if d>1)
            n_named_once = sum(1 for n,d in indeg.items() if d==1)
            n_not_named = a.get("n_periphery",0) + a.get("n_survey_isolated",0)
            st.markdown(insight_card(
                "DENSITY OF WEAK TIES", INDIGO,
                "Who is known, and by how many",
                f"{n_named_multi} named by 2+",
                INDIGO,
                f"{n_named_multi} people are named by multiple respondents — a sign of shared recognition. {n_named_once} appear in only one person's network. {n_not_named} are not named by anyone else, suggesting potential gaps in mutual visibility.",
                "Key attribute: density of weak ties"
            ), unsafe_allow_html=True)

            # Mutual Obligation
            st.markdown(f'<div style="font-family:Barlow Condensed,sans-serif;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:{TEAL};border-bottom:2px solid {TEAL};padding-bottom:6px;margin-bottom:12px;margin-top:20px;">4 · Mutual Obligation</div>', unsafe_allow_html=True)
            trust_color = GREEN if trust_pct>=60 else (AMBER if trust_pct>=40 else TERRA)
            st.markdown(insight_card(
                "TRUST", trust_color,
                "Care that holds when something tests it",
                f"{trust_pct}% rate trust highly",
                trust_color,
                f"Trust under stress is the diagnostic attribute of mutual obligation. {f'{trust_low}% of connections score low on trust — these are the relationships where mutual obligation is most fragile.' if trust_low>15 else 'The network shows a relatively strong trust foundation.'} Reciprocity: {pct_balanced}% of connections are described as roughly balanced.",
                "Key attribute: trust under stress"
            ), unsafe_allow_html=True)
            st.markdown(insight_card(
                "ENERGY & CREATIVITY", TEAL,
                "Generativity across connections",
                f"{energy_pct}% / {a.get('pct_creativity',0)}%",
                TEAL,
                f"Energy ({energy_pct}% high) and Creativity ({a.get('pct_creativity',0)}% high) track the aliveness of relationships. Where these are low, mutual obligation may exist in name but not in felt experience.",
                "Key attribute: reciprocity of care"
            ), unsafe_allow_html=True)

        with col_r:
            # Institutional Anchors
            st.markdown(f'<div style="font-family:Barlow Condensed,sans-serif;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:{AMBER};border-bottom:2px solid {AMBER};padding-bottom:6px;margin-bottom:12px;">6 · Institutional Anchors</div>', unsafe_allow_html=True)
            if "Sector" in ego_df.columns:
                sectors = ego_df["Sector"].dropna().value_counts()
                top_sector = sectors.index[0] if len(sectors) else "N/A"
                n_sectors = len(sectors)
                st.markdown(insight_card(
                    "DISTRIBUTION ACROSS FUNCTION", AMBER,
                    "Coverage across the ecosystem",
                    f"{n_sectors} sector{'s' if n_sectors!=1 else ''}",
                    AMBER,
                    f"The network spans {n_sectors} organizational sectors, led by {top_sector}. The framework asks whether institutional coverage spans the full range of functions — this is where gaps in sector representation become visible.",
                    "Key attribute: distribution across function"
                ), unsafe_allow_html=True)

            # Network structure
            st.markdown(f'<div style="font-family:Barlow Condensed,sans-serif;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:{GREEN};border-bottom:2px solid {GREEN};padding-bottom:6px;margin-bottom:12px;margin-top:20px;">Structure · Depth · Geography</div>', unsafe_allow_html=True)
            depth_color = GREEN if pct_deep>=50 else AMBER
            st.markdown(insight_card(
                "DEPTH OF CONNECTION", depth_color,
                "From awareness to collaboration",
                f"{pct_deep}% at cooperation+",
                depth_color,
                f"The framework distinguishes presence from grounding. Depth is the network equivalent: {pct_deep}% of connections reach Cooperation or Collaboration, meaning they involve real working exchange — not just awareness of one another.",
                "Framework: activated vs. potential"
            ), unsafe_allow_html=True)

            if n_communities > 1:
                st.markdown(insight_card(
                    "CLUSTERS", INDIGO,
                    "Distinct groups within the network",
                    f"{n_communities} clusters",
                    INDIGO,
                    f"The network contains {n_communities} distinct clusters — groups that interact more within themselves than across. The framework's boundary question applies here: whose 'we' is each cluster built around, and who sits between them?",
                    "Framework: boundary question"
                ), unsafe_allow_html=True)

            if n_geographies > 1:
                geo_color = AMBER if n_geographies > 3 else GREEN
                st.markdown(insight_card(
                    "GEOGRAPHY", geo_color,
                    "Where the network lives",
                    f"{n_geographies} area{'s' if n_geographies!=1 else ''}",
                    geo_color,
                    f"The network spans {n_geographies} geographic areas. The framework notes that geography excludes, relationship includes — the asymmetry has to be held with intentionality, especially for a distributed network.",
                    "Framework: geography excludes"
                ), unsafe_allow_html=True)

        # ── What to watch ──
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f'<div class="eyebrow" style="color:{TERRA};">Signals worth watching</div>', unsafe_allow_html=True)
        watch_items = []
        if trust_low > 15:
            watch_items.append(("Trust gap", f"{trust_low}% of connections score low on trust.", TERRA))
        if a.get("n_survey_isolated",0) > 0:
            watch_items.append(("Isolation", f"{a.get('n_survey_isolated',0)} survey taker(s) not named by anyone else — the loneliness the proposal flags.", TERRA))
        if a.get("top_bridge"):
            watch_items.append(("Bridge concentration", f"One person carries disproportionate connective load (betweenness {a.get('top_bridge_score',0):.3f}). Structural asset and structural vulnerability.", AMBER))
        if energy_low > 25:
            watch_items.append(("Low-energy connections", f"{energy_low}% of connections feel low-energy — worth asking which ones, and why.", AMBER))
        if not watch_items:
            watch_items.append(("No major signals", "The network looks healthy across the key dimensions.", GREEN))

        w_cols = st.columns(len(watch_items)) if len(watch_items) <= 3 else st.columns(3)
        for i,(label,text,color) in enumerate(watch_items):
            with w_cols[i % len(w_cols)]:
                st.markdown(f'<div style="background:{SURFACE2};border-top:3px solid {color};padding:16px;border-radius:0 0 6px 6px;"><div style="font-family:IBM Plex Mono,monospace;font-size:10px;text-transform:uppercase;letter-spacing:0.1em;color:{color};margin-bottom:6px;">{label}</div><div style="font-family:IBM Plex Sans,sans-serif;font-size:13px;color:{TEXT2};line-height:1.6;">{text}</div></div>', unsafe_allow_html=True)

# ═══ TAB 5: ROOM VIEW ════════════════════════════════════════════════════════
with tab5:
    st.markdown(f'<div class="eyebrow" style="color:{TEAL};">Room-ready insights · Activity 5 — Who is in the ecosystem?</div>',unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:13px;color:{TEXT2};margin-bottom:16px;">Full-screen anonymized document — no names, no individual data. Project this in the room during Activity 5. Download from the sidebar.</p>',unsafe_allow_html=True)
    if not a:
        st.info("Process data to generate the room view.")
    else:
        html_preview = build_insights_html(a, ego_df, edge_df)
        st.components.v1.html(html_preview, height=720, scrolling=True)
